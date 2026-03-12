import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..models.funding import FundingSearchResult


def generate_pdf(results: list[FundingSearchResult], query: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    # Titolo
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=20)
    elements.append(Paragraph("GrantFinder AI — Report Finanziamenti", title_style))
    elements.append(Paragraph(f"<b>Ricerca:</b> {query[:200]}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Data:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    for i, r in enumerate(results, 1):
        elements.append(Paragraph(f"<b>#{i} — {r.title}</b>", styles["Heading2"]))
        data = [
            ["Programma", r.program or "N/D"],
            ["Importo", f"€{r.min_grant:,.0f} - €{r.max_grant:,.0f}" if r.min_grant and r.max_grant else "N/D"],
            ["Cofinanziamento", f"{r.funding_percentage}%" if r.funding_percentage else "N/D"],
            ["Scadenza", r.deadline.strftime("%d/%m/%Y") if r.deadline else "N/D"],
            ["Score", f"{r.relevance_score:.1%}"],
        ]
        t = Table(data, colWidths=[120, 350])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.95)),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        if r.why_relevant:
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"<i>Perche' rilevante: {r.why_relevant}</i>", styles["Normal"]))
        if r.url:
            elements.append(Paragraph(f'<link href="{r.url}">Link ufficiale</link>', styles["Normal"]))
        elements.append(Spacer(1, 16))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel(results: list[FundingSearchResult], query: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finanziamenti"

    headers = ["#", "Titolo", "Programma", "Ente", "Importo Min", "Importo Max",
               "% Cofin.", "Scadenza", "Tipo", "Ambito", "Score", "Link"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.title)
        ws.cell(row=row, column=3, value=r.program or "")
        ws.cell(row=row, column=4, value=r.organization or "")
        ws.cell(row=row, column=5, value=float(r.min_grant) if r.min_grant else None)
        ws.cell(row=row, column=6, value=float(r.max_grant) if r.max_grant else None)
        ws.cell(row=row, column=7, value=float(r.funding_percentage) if r.funding_percentage else None)
        ws.cell(row=row, column=8, value=r.deadline.strftime("%d/%m/%Y") if r.deadline else "")
        ws.cell(row=row, column=9, value=r.funding_type or "")
        ws.cell(row=row, column=10, value=", ".join(r.geographic_scope) if r.geographic_scope else "")
        ws.cell(row=row, column=11, value=round(r.relevance_score, 3))
        ws.cell(row=row, column=12, value=r.url or "")

    for col in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
